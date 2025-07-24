import subprocess
import sys
import pandas as pd
from openpyxl import load_workbook

TEMP_SHIFT_PATH = "temp_shift_final.csv"
TEMPLATE_PATH = "shift_template.xlsx"
OUTPUT_PATH = "shift_output.xlsx"

# 1. optimize_1.py を実行して temp_shift_output_1.csv を生成
print("=== Step 1: Strict1~3の最適化 ===")
ret1 = subprocess.run([sys.executable, "optimize_1.py"])
if ret1.returncode != 0:
    print("optimize_1.py の実行に失敗しました")
    sys.exit(1)

# 2. optimize_2.py を実行して、temp_shift_output_1.csv を使って次の最適化
print("=== Step 2: 追加最適化 ===")
ret2 = subprocess.run([sys.executable, "optimize_2.py"])
if ret2.returncode != 0:
    print("optimize_2.py の実行に失敗しました")
    sys.exit(1)
# CSV読込
df = pd.read_csv(TEMP_SHIFT_PATH, index_col=0)

# Excelテンプレートをロード
wb = load_workbook(TEMPLATE_PATH)
sheet = wb['シフト表']

# テンプレートの看護師リスト(A6:A19など)を取得
nurse_rows = list(range(6, 6 + len(df.index)))
nurse_names_in_excel = [sheet.cell(row=r, column=1).value for r in nurse_rows]

# 日付の列(C列:3番目以降)
col_offset = 3
date_cols = df.columns.tolist()[:31]  # 31日分のみ

# データを書き込み
for i, nurse in enumerate(df.index):
    if nurse not in nurse_names_in_excel:
        continue
    row_idx = nurse_names_in_excel.index(nurse) + 6
    for j, day_col in enumerate(date_cols):
        shift = df.at[nurse, day_col]
        sheet.cell(row=row_idx, column=col_offset + j, value=shift)

wb.save(OUTPUT_PATH)
print(f"✅ シフト表を {OUTPUT_PATH} に保存しました。")


print("=== 完了 ===")