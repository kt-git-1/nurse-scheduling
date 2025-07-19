from ortools.sat.python import cp_model
import pandas as pd
import calendar
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook

# === 設定 ===
YEAR = 2025
MONTH = 8
DAYS_IN_MONTH = 31
NURSES = ['久保', '三好', '御書', '板川', '友枝', '樋渡', '中山', '川原田', '奥平', '前野', '森園', '小嶋', '田浦', '久保（千）']
SHIFT_TYPES = ['休', '夜', '早', '残', '〇', '1', '2', '3', '4', '×', '/休', '休/']
HOLIDAY_WEEKDAYS = ['木', '日']
HALF_HOLIDAY_WEEKDAYS = ['土']

# === 希望種別の変換 ===
希望種別マップ = {
    '①': ['休'],
    '②': ['休', '×'],
    '③': ['/休'],
    '④': ['休/'],
    '⑤': ['休/', '×'],
}

# === ファイルパス ===
INPUT_CSV = '/Users/kaito_taniguchi/workspace/nurse-shift/kibou_input.csv'
TEMPLATE_PATH = '/Users/kaito_taniguchi/workspace/nurse-shift/shift_template.xlsx'
OUTPUT_PATH = '/Users/kaito_taniguchi/workspace/nurse-shift/shift_output.xlsx'

# === 日付と曜日を生成 ===
start_date = datetime(YEAR, MONTH - 1, 21)
dates = [start_date + timedelta(days=i) for i in range(DAYS_IN_MONTH)]
date_labels = [f"{d.day}" for d in dates]
weekdays = [calendar.day_name[d.weekday()][0] for d in dates]

# === OR-Tools モデル定義 ===
model = cp_model.CpModel()
x = {}

for n in NURSES:
    for d in range(DAYS_IN_MONTH):
        for s in SHIFT_TYPES:
            x[n, d, s] = model.NewBoolVar(f'x_{n}_{d}_{s}')

# === 制約 ===
for d in range(DAYS_IN_MONTH):
    model.AddExactlyOne(x[n, d, '夜'] for n in NURSES)

for n in ['板川', '三好', '御書']:
    for d in range(DAYS_IN_MONTH):
        model.Add(x[n, d, '夜'] == 0)
        if n == '御書':
            model.Add(x[n, d, '早'] == 0)
            model.Add(x[n, d, '残'] == 0)

for n in NURSES:
    for d in range(DAYS_IN_MONTH):
        model.AddAtMostOne(x[n, d, s] for s in SHIFT_TYPES)

for n in NURSES:
    for d in range(DAYS_IN_MONTH - 1):
        model.AddImplication(x[n, d, '夜'], x[n, d + 1, '×'])

# === 希望休のソフト制約 ===
penalties = []
希望df = pd.read_csv(INPUT_CSV, encoding='utf-8')

名前リスト = 希望df['日付'].tolist()
日付列 = [col for col in 希望df.columns if str(col).isdigit()]

for i, nurse in enumerate(名前リスト):
    for col in 日付列:
        typ = str(希望df.loc[i, col]).strip() if pd.notna(希望df.loc[i, col]) else ''
        if typ in 希望種別マップ:
            day = int(col) - 1  # 0-indexed day
            allowed = 希望種別マップ[typ]
            violated = model.NewBoolVar(f'violated_{nurse}_{day}')
            model.AddBoolOr([x[nurse, day, s] for s in allowed]).OnlyEnforceIf(violated.Not())
            penalties.append(violated)

model.Minimize(sum(penalties))

# === 求解 ===
solver = cp_model.CpSolver()
status = solver.Solve(model)

if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
    print("✅ 最適化成功：テンプレートに書き込みます...")

    # テンプレートを開く
    wb = load_workbook(TEMPLATE_PATH)
    sheet = wb['シフト表']
    if sheet is None:
        raise ValueError("No active worksheet found in the template.")

    # 看護師名をテンプレートから取得（A6〜A19）
    nurses_in_sheet = [sheet.cell(row=i, column=1).value for i in range(6, 20)]

    # 日付列はC=3からAG=33（31列）
    for row_index, nurse in enumerate(nurses_in_sheet):
        if nurse not in NURSES:
            continue
        for col_index in range(DAYS_IN_MONTH):
            shift = ''
            for s in SHIFT_TYPES:
                if solver.Value(x[nurse, col_index, s]):
                    shift = s
                    break
            sheet.cell(row=6 + row_index, column=3 + col_index, value=shift)

    # 保存
    wb.save(OUTPUT_PATH)
    print(f"✅ シフト表を {OUTPUT_PATH} に保存しました。")
else:
    print("❌ 解が見つかりませんでした。")
